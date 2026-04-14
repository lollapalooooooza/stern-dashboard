#!/usr/bin/env python3

import argparse
import ast
import base64
import gzip
import json
import math
import operator
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import yfinance as yf
from openpyxl import load_workbook


THEME_NAME_MAP = {
    "AI Infra": "AI-Industrial",
    "Silver Economy": "Silver",
    "Silver": "Silver",
    "Batteries": "Battery",
    "Waste Management": "Waste",
    "Digital Finance": "Payments",
    "Legacy": "Legacy Software",
    "Water PFAs": "Water PFAS",
    "PFAs": "Water PFAS",
    "Securities": "Security",
}

FALLBACK_COMPANY_NAMES = {
    "AON": "Aon plc",
    "BAC": "Bank of America",
    "BANC": "Banc of California",
    "C": "Citigroup",
    "CB": "Chubb",
    "DHT": "DHT Holdings",
    "EWBC": "East West Bancorp",
    "FHN": "First Horizon",
    "FRO": "Frontline plc",
    "GS": "Goldman Sachs",
    "JPM": "JPMorgan Chase",
    "LMT": "Lockheed Martin",
    "MS": "Morgan Stanley",
    "MTB": "M&T Bank",
    "PNFP": "Pinnacle Financial Partners",
    "WFC": "Wells Fargo",
}

DB_ROW_COLUMNS = [
    "id",
    "ticker",
    "company",
    "theme",
    "subTheme",
    "buyPrice",
    "currentPrice",
    "entryDate",
    "exitDate",
    "shares",
    "benchmarkWeight",
    "stopLossPct",
    "status",
    "notes",
    "marketBeta",
    "valueBeta",
    "momentumBeta",
    "weeklyReturn",
    "currentValue",
    "pnlFromExcel",
    "sellPrice",
    "costBasis",
    "sellTotal",
    "realizedPnl",
    "realizedPnlPct",
]

TRACKER_DATE_CUTOFF = date(2026, 3, 31)
SAFE_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
SAFE_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def parse_args():
    parser = argparse.ArgumentParser(description="Rebuild workbook-backed thematic seed data")
    parser.add_argument("--positions", required=True, help="Path to current positions workbook")
    parser.add_argument("--tracker", required=True, help="Path to entry/exit tracker workbook")
    parser.add_argument("--balance", help="Path to portfolio balance workbook")
    parser.add_argument(
        "--benchmark-weight",
        type=float,
        default=0.25,
        help="Target SPY benchmark weight as a percent of total portfolio value",
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parents[1] / "lib" / "thematicWorkbookData.js"),
        help="Path to generated JS module",
    )
    return parser.parse_args()


def theme_name(raw):
    return THEME_NAME_MAP.get(raw, raw)


def date_key(value):
    if hasattr(value, "date"):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def normalize_ledger_date(value):
    if not hasattr(value, "date") and not isinstance(value, date):
        return None
    normalized = value.date() if hasattr(value, "date") else value
    while normalized > TRACKER_DATE_CUTOFF:
        normalized = normalized.replace(year=normalized.year - 1)
    return normalized.isoformat()


def normalize_balance_date(value, previous_date=None):
    if not hasattr(value, "date") and not isinstance(value, date):
        return None
    normalized = value.date() if hasattr(value, "date") else value
    while normalized > TRACKER_DATE_CUTOFF:
        normalized = normalized.replace(year=normalized.year - 1)
    while previous_date and normalized < previous_date and (previous_date - normalized).days > 180:
        normalized = normalized.replace(year=normalized.year + 1)
    return normalized.isoformat()


def as_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if not stripped:
            return None
        if stripped.lower() in {"nan", "inf", "+inf", "-inf"}:
            return None
        if stripped.startswith("="):
            stripped = stripped[1:]
        try:
            numeric = float(stripped)
            return numeric if math.isfinite(numeric) else None
        except (TypeError, ValueError):
            try:
                parsed = ast.parse(stripped, mode="eval")
                numeric = float(eval_numeric_formula(parsed.body))
                return numeric if math.isfinite(numeric) else None
            except (SyntaxError, TypeError, ValueError, ZeroDivisionError):
                return None
    try:
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    except (TypeError, ValueError):
        return None


def eval_numeric_formula(node):
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.UnaryOp) and type(node.op) in SAFE_UNARY_OPERATORS:
        return SAFE_UNARY_OPERATORS[type(node.op)](eval_numeric_formula(node.operand))
    if isinstance(node, ast.BinOp) and type(node.op) in SAFE_BINARY_OPERATORS:
        return SAFE_BINARY_OPERATORS[type(node.op)](eval_numeric_formula(node.left), eval_numeric_formula(node.right))
    raise ValueError(f"Unsupported formula node: {ast.dump(node)}")


def load_existing_seed(seed_path):
    text = seed_path.read_text()
    b64 = "".join(re.findall(r'"([A-Za-z0-9+/=]+)"', text))
    rows = json.loads(gzip.decompress(base64.b64decode(b64)).decode())
    active_by_ticker = {}
    any_by_ticker = {}
    for row in rows:
        meta = {
            "company": row[2],
            "theme": row[3],
            "subTheme": row[4],
            "stopLossPct": float(row[11] or 0.1),
            "notes": row[13] or "",
            "marketBeta": float(row[14] or 1.0),
            "valueBeta": float(row[15] or 0.0),
            "momentumBeta": float(row[16] or 0.0),
            "benchmarkWeight": float(row[10] or 0.0),
            "buyPrice": float(row[5] or 0.0),
        }
        any_by_ticker.setdefault(row[1], meta)
        if row[12] == "active":
            active_by_ticker.setdefault(row[1], meta)
    return active_by_ticker, any_by_ticker


def load_existing_workbook_payload(module_path):
    text = module_path.read_text()
    match = re.search(r'const THEMATIC_WORKBOOK_B64 = \[(.*?)\]\.join\(\"\"\);', text, re.S)
    if not match:
        raise ValueError(f"Could not decode workbook module: {module_path}")
    b64 = "".join(re.findall(r'"([A-Za-z0-9+/=]+)"', match.group(1)))
    return json.loads(gzip.decompress(base64.b64decode(b64)).decode())


def build_workbook_meta(payload):
    active_by_ticker = {}
    any_by_ticker = {}
    row_by_ticker = {}
    for row in payload["activeRows"] + payload["exitedRows"]:
        meta = {
            "company": row[2],
            "theme": row[3],
            "subTheme": row[4],
            "buyPrice": float(row[5] or 0),
            "currentPrice": float(row[6] or 0),
            "shares": float(row[9] or 0),
            "benchmarkWeight": float(row[10] or 0),
            "stopLossPct": float(row[11] or 0.1),
            "notes": row[13] or "",
            "marketBeta": float(row[14] or 1.0),
            "valueBeta": float(row[15] or 0.0),
            "momentumBeta": float(row[16] or 0.0),
            "currentValue": float(row[18] or 0),
            "costBasis": float(row[21] or 0),
            "status": row[12],
        }
        any_by_ticker[row[1]] = meta
        row_by_ticker[row[1]] = row
        if row[12] == "active":
            active_by_ticker[row[1]] = meta
    return active_by_ticker, any_by_ticker, row_by_ticker, payload["summary"], payload["weeklyHistory"], payload["dailyHistory"]


def load_positions(positions_path):
    wb = load_workbook(positions_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    active_rows = []
    summary = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = list(values[:6]) + [None] * max(0, 6 - len(values[:6]))
        symbol, subtheme, weight, current_value, pnl, pnl_pct = row[:6]
        if symbol and subtheme:
            active_rows.append(
                {
                    "ticker": symbol,
                    "theme": theme_name(subtheme),
                    "weight": float(weight or 0),
                    "currentValue": float(current_value or 0),
                    "pnl": float(pnl or 0),
                    "pnlPct": float(pnl_pct or 0),
                }
            )
        elif symbol and subtheme is None:
            summary[symbol] = float(current_value or 0)
    return active_rows, summary


def load_tracker(tracker_path):
    wb = load_workbook(tracker_path, data_only=False)

    ledger = []
    ws_ledger = wb["Sheet1"]
    current_theme = None
    last_slot_dates = {2: None, 6: None, 10: None, 16: None, 20: None}
    for idx, row in enumerate(ws_ledger.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            current_theme = row[0]
            last_slot_dates = {2: None, 6: None, 10: None, 16: None, 20: None}
        ticker = row[1]
        if not ticker:
            continue

        entries = []
        for base in (2, 6, 10):
            entry_date, quantity, average_cost, total = row[base : base + 4]
            normalized_date = normalize_ledger_date(entry_date)
            if normalized_date:
                last_slot_dates[base] = normalized_date
            qty_value = as_float(quantity)
            cost_value = as_float(average_cost)
            total_value = as_float(total)
            if qty_value or cost_value or total_value:
                entries.append(
                    {
                        "date": normalized_date or last_slot_dates[base],
                        "qty": qty_value or 0,
                        "cost": cost_value or 0,
                        "total": total_value or 0,
                    }
                )

        exits = []
        for base in (16, 20):
            exit_date, quantity, average_price, total = row[base : base + 4]
            normalized_date = normalize_ledger_date(exit_date)
            if normalized_date:
                last_slot_dates[base] = normalized_date
            qty_value = as_float(quantity)
            price_value = as_float(average_price)
            total_value = as_float(total)
            if qty_value or price_value or total_value:
                exits.append(
                    {
                        "date": normalized_date or last_slot_dates[base],
                        "qty": qty_value or 0,
                        "price": price_value or 0,
                        "total": total_value or 0,
                    }
                )

        entry_qty = sum(entry["qty"] for entry in entries)
        exit_qty = sum(exit_row["qty"] for exit_row in exits)
        ledger.append(
            {
                "rowIdx": idx,
                "ticker": ticker,
                "theme": theme_name(current_theme),
                "rawTheme": current_theme,
                "entries": entries,
                "exits": exits,
                "entryQty": entry_qty,
                "exitQty": exit_qty,
                "remainingQty": entry_qty - exit_qty,
                "allFlag": row[15],
            }
        )

    ws_hist = wb["Historical Data"]
    headers = [cell.value for cell in ws_hist[1]]
    raw_by_date = {}
    for row in ws_hist.iter_rows(min_row=2, values_only=True):
        row_date = date_key(row[0])
        if not row_date:
            continue
        raw_by_date[row_date] = {
            headers[index]: as_float(row[index]) for index in range(1, len(headers)) if headers[index]
        }

    tickers = headers[1:]
    filled_by_date = {}
    last_valid = {ticker: None for ticker in tickers if ticker}
    for row_date in sorted(raw_by_date):
        current = {}
        for ticker in tickers:
            if not ticker:
                continue
            value = raw_by_date[row_date].get(ticker)
            if value is not None and not math.isnan(value):
                last_valid[ticker] = float(value)
            current[ticker] = last_valid[ticker]
        filled_by_date[row_date] = current

    return ledger, filled_by_date


def load_balance_tracker(balance_path):
    wb = load_workbook(balance_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    previous_date = None

    for values in ws.iter_rows(min_row=2, values_only=True):
        raw_date, total_balance, monthly_return, since_start = values[:4]
        normalized_date = normalize_balance_date(raw_date, previous_date)
        if not normalized_date or total_balance is None:
            continue
        rows.append(
            {
                "date": normalized_date,
                "portfolioValue": float(total_balance or 0),
                "monthlyReturn": as_float(monthly_return),
                "sinceStart": as_float(since_start),
            }
        )
        previous_date = date.fromisoformat(normalized_date)

    return rows


def download_spy_history(start_date, end_date):
    history = yf.download("SPY", start=start_date, end=end_date, auto_adjust=False, progress=False)
    close_series = history["Close"]["SPY"] if getattr(history.columns, "nlevels", 1) > 1 else history["Close"]
    return {timestamp.strftime("%Y-%m-%d"): float(value) for timestamp, value in close_series.items() if not math.isnan(float(value))}


def download_latest_closes(tickers):
    tickers = [ticker for ticker in tickers if ticker]
    if not tickers:
        return {}
    history = yf.download(
        tickers,
        period="1mo",
        auto_adjust=False,
        progress=False,
        group_by="ticker",
        threads=False,
    )
    if history is None or getattr(history, "empty", False):
        return {}

    prices = {}
    columns = getattr(history, "columns", None)
    nlevels = getattr(columns, "nlevels", 1)
    if len(tickers) == 1 and nlevels == 1:
        close_series = history["Close"].dropna()
        if not close_series.empty:
            prices[tickers[0]] = float(close_series.iloc[-1])
        return prices

    if nlevels > 1:
        for ticker in tickers:
            if ticker not in history.columns.get_level_values(0):
                continue
            close_series = history[ticker]["Close"].dropna()
            if not close_series.empty:
                prices[ticker] = float(close_series.iloc[-1])
    return prices


def date_from_filename(path):
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if not match:
        return None
    month, day, year = map(int, match.groups())
    return date(year, month, day).isoformat()


def nearest_available_date(sorted_dates, target_date):
    candidates = [row_date for row_date in sorted_dates if row_date <= target_date]
    return candidates[-1] if candidates else sorted_dates[0]


def trading_sessions_back(sorted_dates, target_date, sessions):
    index = sorted_dates.index(target_date)
    return sorted_dates[max(0, index - sessions)]


def find_metadata(ticker, active_meta, any_meta):
    return active_meta.get(ticker) or any_meta.get(ticker) or {
        "company": FALLBACK_COMPANY_NAMES.get(ticker, ticker),
        "theme": "",
        "subTheme": "",
        "stopLossPct": 0.1,
        "notes": "",
        "marketBeta": 1.0,
        "valueBeta": 0.0,
        "momentumBeta": 0.0,
        "benchmarkWeight": 0.0,
        "buyPrice": 0.0,
    }


def entry_total(entry):
    return entry["total"] or (entry["qty"] * entry["cost"])


def exit_total(exit_row):
    return exit_row["total"] or (exit_row["qty"] * exit_row["price"])


def display_company_name(ticker, meta):
    company = meta.get("company") if meta else None
    if company and company != ticker:
        return company
    return FALLBACK_COMPANY_NAMES.get(ticker, ticker)


def normalize_theme(value):
    return theme_name(value or "").strip().lower()


def match_active_ledger_rows(active_rows, ledger_rows):
    active_ledger_by_ticker = defaultdict(list)
    for row in ledger_rows:
        if row["remainingQty"] > 1e-8:
            active_ledger_by_ticker[row["ticker"]].append(row)

    used = set()
    for row in active_rows:
        candidates = active_ledger_by_ticker.get(row["ticker"], [])
        matched = None
        for candidate in candidates:
            if candidate["rowIdx"] in used:
                continue
            if normalize_theme(candidate["theme"]) == normalize_theme(row["theme"]):
                matched = candidate
                break
        if matched is None:
            for candidate in candidates:
                if candidate["rowIdx"] not in used:
                    matched = candidate
                    break
        if matched is None and candidates:
            matched = candidates[0]
        if matched is not None:
            used.add(matched["rowIdx"])
        row["matchedLedger"] = matched


def build_active_rows(
    active_rows,
    active_meta,
    any_meta,
    price_rows,
    reference_date,
    valuation_date,
    previous_week_date,
    supplemental_prices,
):
    current_shares_by_ticker = defaultdict(float)
    for row in active_rows:
        meta = find_metadata(row["ticker"], active_meta, any_meta)
        reference_price = (
            price_rows.get(reference_date, {}).get(row["ticker"])
            or supplemental_prices.get(row["ticker"])
            or meta.get("currentPrice")
            or meta.get("buyPrice")
        )
        current_price = (
            price_rows.get(valuation_date, {}).get(row["ticker"])
            or supplemental_prices.get(row["ticker"])
            or meta.get("currentPrice")
            or reference_price
        )
        if not reference_price or reference_price <= 0:
            raise ValueError(f"Missing reference price for {row['ticker']}")
        if not current_price or current_price <= 0:
            raise ValueError(f"Missing valuation price for {row['ticker']}")
        previous_price = price_rows[previous_week_date].get(row["ticker"]) or current_price
        row["shares"] = row["currentValue"] / float(current_price)
        matched = row.get("matchedLedger")
        if row["pnl"]:
            row["costBasis"] = row["currentValue"] - row["pnl"]
        elif matched and matched["remainingQty"] > 1e-8:
            total_entry_cost = sum(entry_total(entry) for entry in matched["entries"])
            sold_cost_basis = total_entry_cost * (matched["exitQty"] / matched["entryQty"]) if matched["entryQty"] else 0
            remaining_cost_basis = max(total_entry_cost - sold_cost_basis, 0)
            scale = row["shares"] / matched["remainingQty"] if matched["remainingQty"] else 1.0
            row["costBasis"] = remaining_cost_basis * scale
        else:
            fallback_buy_price = meta["buyPrice"] if meta["buyPrice"] > 0 else float(current_price)
            row["costBasis"] = row["shares"] * fallback_buy_price
            if row["costBasis"] <= 0:
                row["costBasis"] = row["currentValue"]
        row["buyPrice"] = row["costBasis"] / row["shares"] if row["shares"] else float(current_price)
        row["currentPrice"] = float(current_price)
        row["currentValue"] = float(row["currentValue"])
        row["pnl"] = row["currentValue"] - row["costBasis"]
        row["weeklyReturn"] = (row["currentPrice"] / previous_price) - 1 if previous_price else 0
        current_shares_by_ticker[row["ticker"]] += row["shares"]

    db_rows = []
    for index, row in enumerate(active_rows, start=1):
        meta = find_metadata(row["ticker"], active_meta, any_meta)
        matched = row.get("matchedLedger")
        entry_date = ""
        if matched:
            valid_dates = [entry["date"] for entry in matched["entries"] if entry["date"]]
            entry_date = valid_dates[0] if valid_dates else ""
        db_rows.append(
            [
                f"active-{index}",
                row["ticker"],
                display_company_name(row["ticker"], meta),
                row["theme"],
                meta["subTheme"] or row["theme"],
                row["buyPrice"],
                row["currentPrice"],
                entry_date,
                "",
                row["shares"],
                meta["benchmarkWeight"],
                meta["stopLossPct"],
                "active",
                meta["notes"] or f"{row['theme']} snapshot holding",
                meta["marketBeta"],
                meta["valueBeta"],
                meta["momentumBeta"],
                row["weeklyReturn"],
                row["currentValue"],
                row["pnl"],
                0,
                0,
                0,
                0,
                0,
            ]
        )
    return db_rows, current_shares_by_ticker


def build_benchmark_row(
    active_meta,
    any_meta,
    reference_date,
    valuation_date,
    reference_spy_price,
    valuation_spy_price,
    benchmark_value,
    previous_spy_price=None,
):
    meta = find_metadata("SPY", active_meta, any_meta)
    valuation_price = valuation_spy_price or reference_spy_price
    shares = benchmark_value / valuation_price if valuation_price else 0
    current_value = benchmark_value
    pnl = current_value - benchmark_value
    weekly_return = (valuation_price / previous_spy_price) - 1 if previous_spy_price else 0
    return [
        "benchmark-spy",
        "SPY",
        display_company_name("SPY", meta) if meta else "SPDR S&P 500 ETF Trust",
        "Benchmark",
        meta["subTheme"] or "Benchmark",
        reference_spy_price or valuation_price,
        valuation_price,
        reference_date,
        "",
        shares,
        meta["benchmarkWeight"],
        meta["stopLossPct"],
        "active",
        meta["notes"] or "S&P 500 core",
        1.0,
        meta["valueBeta"],
        meta["momentumBeta"],
        weekly_return,
        current_value,
        pnl,
        0,
        0,
        0,
        0,
        0,
    ], shares


def build_exited_rows(ledger_rows, scale_by_ticker, active_meta, any_meta):
    db_rows = []
    counter = 1
    for row in ledger_rows:
        if row["exitQty"] <= 1e-8:
            continue

        scale = scale_by_ticker.get(row["ticker"], 1.0) if row["remainingQty"] > 1e-8 else 1.0
        sold_qty = row["exitQty"] * scale
        sold_total = sum(exit_total(exit_row) for exit_row in row["exits"]) * scale
        total_entry_cost = sum(entry_total(entry) for entry in row["entries"]) * scale
        total_entry_qty = row["entryQty"] * scale
        sold_cost_basis = total_entry_cost * (sold_qty / total_entry_qty) if total_entry_qty else 0
        realized_pnl = sold_total - sold_cost_basis
        realized_pnl_pct = realized_pnl / sold_cost_basis if sold_cost_basis else 0
        sell_price = sold_total / sold_qty if sold_qty else 0
        buy_price = sold_cost_basis / sold_qty if sold_qty else 0
        exit_dates = [exit_row["date"] for exit_row in row["exits"] if exit_row["date"]]
        entry_dates = [entry["date"] for entry in row["entries"] if entry["date"]]
        meta = find_metadata(row["ticker"], active_meta, any_meta)
        db_rows.append(
            [
                f"exited-{counter}",
                row["ticker"],
                display_company_name(row["ticker"], meta),
                row["theme"],
                meta["subTheme"] or row["theme"],
                buy_price or meta["buyPrice"],
                sell_price,
                entry_dates[0] if entry_dates else "",
                exit_dates[-1] if exit_dates else "",
                sold_qty,
                meta["benchmarkWeight"],
                meta["stopLossPct"],
                "exited",
                meta["notes"] or f"{row['theme']} exited position",
                meta["marketBeta"],
                meta["valueBeta"],
                meta["momentumBeta"],
                0,
                sold_total,
                realized_pnl,
                sell_price,
                sold_cost_basis,
                sold_total,
                realized_pnl,
                realized_pnl_pct,
            ]
        )
        counter += 1
    return db_rows


def build_scale_map(ledger_rows, current_shares_by_ticker):
    remaining_by_ticker = defaultdict(float)
    for row in ledger_rows:
        if row["remainingQty"] > 1e-8:
            remaining_by_ticker[row["ticker"]] += row["remainingQty"]

    scale = {}
    for ticker, remaining_qty in remaining_by_ticker.items():
        if remaining_qty > 0:
            scale[ticker] = current_shares_by_ticker[ticker] / remaining_qty
    return scale


def build_daily_history(balance_rows, spy_rows):
    spy_dates = sorted(spy_rows)
    latest_spy = spy_rows[spy_dates[0]] if spy_dates else None
    spy_index = 0
    previous_portfolio_value = None
    previous_spy_price = None
    benchmark_nav = 1.0
    daily_rows = []

    for row in balance_rows:
        row_date = row["date"]
        while spy_index < len(spy_dates) and spy_dates[spy_index] <= row_date:
            latest_spy = spy_rows[spy_dates[spy_index]]
            spy_index += 1
        if latest_spy is None:
            continue

        portfolio_value = row["portfolioValue"]
        if previous_portfolio_value is None:
            portfolio_return = 0
            benchmark_return = 0
        else:
            portfolio_return = portfolio_value / previous_portfolio_value - 1 if previous_portfolio_value else 0
            benchmark_return = latest_spy / previous_spy_price - 1 if previous_spy_price else 0
            benchmark_nav *= 1 + benchmark_return

        daily_rows.append(
            {
                "date": row_date,
                "portfolioValue": portfolio_value,
                "benchmarkValue": benchmark_nav,
                "portfolioReturn": portfolio_return,
                "benchmarkReturn": benchmark_return,
                "marketContrib": benchmark_return,
                "valueContrib": 0,
                "momentumContrib": 0,
                "alpha": portfolio_return - benchmark_return,
                "sinceStart": row["sinceStart"],
            }
        )
        previous_portfolio_value = portfolio_value
        previous_spy_price = latest_spy

    return daily_rows


def build_weekly_history(daily_rows):
    weekly_groups = []
    current_group = []
    current_key = None

    for row in daily_rows:
        row_date = datetime.fromisoformat(row["date"])
        iso_key = (row_date.isocalendar().year, row_date.isocalendar().week)
        if current_key and iso_key != current_key:
            weekly_groups.append(current_group)
            current_group = []
        current_group.append(row)
        current_key = iso_key

    if current_group:
        weekly_groups.append(current_group)

    weekly_rows = []
    for index, group in enumerate(weekly_groups, start=1):
        portfolio_nav = 1.0
        benchmark_nav = 1.0
        for row in group:
            portfolio_nav *= 1 + row["portfolioReturn"]
            benchmark_nav *= 1 + row["benchmarkReturn"]
        portfolio_return = portfolio_nav - 1
        benchmark_return = benchmark_nav - 1
        weekly_rows.append(
            {
                "week": f"W{index}",
                "date": group[-1]["date"],
                "portfolioReturn": portfolio_return,
                "benchmarkReturn": benchmark_return,
                "marketContrib": benchmark_return,
                "valueContrib": 0,
                "momentumContrib": 0,
                "alpha": portfolio_return - benchmark_return,
            }
        )
    return weekly_rows


def sample_std(values):
    values = [float(value) for value in values]
    if len(values) < 2:
        return 0.0
    mean = sum(values) / len(values)
    variance = sum((value - mean) ** 2 for value in values) / (len(values) - 1)
    return math.sqrt(variance)


def build_payload(active_rows, exited_rows, weekly_history, daily_history, summary):
    return {
        "snapshotDate": summary["snapshotDate"],
        "summary": summary,
        "activeRows": active_rows,
        "exitedRows": exited_rows,
        "weeklyHistory": weekly_history,
        "dailyHistory": daily_history,
    }


def write_module(output_path, payload):
    encoded = base64.b64encode(gzip.compress(json.dumps(payload, separators=(",", ":")).encode())).decode()
    chunks = [encoded[index : index + 120] for index in range(0, len(encoded), 120)]
    lines = [
        'import { gunzipSync } from "node:zlib";',
        "",
        "const THEMATIC_WORKBOOK_B64 = [",
    ]
    lines.extend(f'  "{chunk}",' for chunk in chunks)
    lines.append('].join("");')
    lines.append("")
    lines.append(
        "const THEMATIC_WORKBOOK = JSON.parse(gunzipSync(Buffer.from(THEMATIC_WORKBOOK_B64, \"base64\")).toString(\"utf8\"));"
    )
    lines.append("")
    lines.append("export const THEMATIC_WORKBOOK_SNAPSHOT_DATE = THEMATIC_WORKBOOK.snapshotDate;")
    lines.append("export const THEMATIC_WORKBOOK_SUMMARY = THEMATIC_WORKBOOK.summary;")
    lines.append("export const THEMATIC_WORKBOOK_ACTIVE_ROWS = THEMATIC_WORKBOOK.activeRows;")
    lines.append("export const THEMATIC_WORKBOOK_EXITED_ROWS = THEMATIC_WORKBOOK.exitedRows;")
    lines.append("export const THEMATIC_WORKBOOK_WEEKLY_HISTORY = THEMATIC_WORKBOOK.weeklyHistory;")
    lines.append("export const THEMATIC_WORKBOOK_DAILY_HISTORY = THEMATIC_WORKBOOK.dailyHistory;")
    lines.append("")
    output_path.write_text("\n".join(lines) + "\n")


def main():
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    seed_active_meta, seed_any_meta = load_existing_seed(repo_root / "lib" / "thematicSeed.js")
    output_path = Path(args.output)
    (
        workbook_active_meta,
        workbook_any_meta,
        _workbook_rows,
        existing_summary,
        existing_weekly_history,
        existing_daily_history,
    ) = build_workbook_meta(load_existing_workbook_payload(output_path))
    active_meta = {**seed_active_meta, **workbook_active_meta}
    any_meta = {**seed_any_meta, **workbook_any_meta}

    active_rows, summary_rows = load_positions(Path(args.positions))
    ledger_rows, price_rows = load_tracker(Path(args.tracker))
    balance_rows = load_balance_tracker(Path(args.balance)) if args.balance else []

    sorted_price_dates = sorted(price_rows)
    snapshot_date = date_from_filename(Path(args.positions)) or existing_summary["snapshotDate"]
    reference_date = nearest_available_date(sorted_price_dates, snapshot_date)
    valuation_date = reference_date
    previous_week_date = trading_sessions_back(sorted_price_dates, valuation_date, 5)
    snapshot_dt = date.fromisoformat(snapshot_date)

    match_active_ledger_rows(active_rows, ledger_rows)
    missing_price_tickers = sorted(
        {
            row["ticker"]
            for row in active_rows
            if price_rows.get(reference_date, {}).get(row["ticker"]) is None
        }
    )
    supplemental_prices = download_latest_closes(missing_price_tickers)
    active_db_rows, current_shares_by_ticker = build_active_rows(
        active_rows,
        active_meta,
        any_meta,
        price_rows,
        reference_date,
        valuation_date,
        previous_week_date,
        supplemental_prices,
    )
    scale_by_ticker = build_scale_map(ledger_rows, current_shares_by_ticker)

    stock_value = sum(row[18] for row in active_db_rows)
    stock_weight = sum(row["weight"] for row in active_rows)
    workbook_portfolio_value = as_float(summary_rows.get("Portfolio Value"))
    workbook_benchmark_value = as_float(summary_rows.get("In Benchmark"))
    if stock_weight <= 0 and workbook_portfolio_value is None:
        raise ValueError("Positions workbook did not contain positive active weights")

    if workbook_portfolio_value is not None:
        portfolio_value = float(workbook_portfolio_value)
    else:
        portfolio_value = stock_value / stock_weight

    if workbook_benchmark_value is not None:
        benchmark_value = float(workbook_benchmark_value)
    else:
        benchmark_value = portfolio_value * args.benchmark_weight

    cash_value = portfolio_value - stock_value - benchmark_value
    if cash_value < -1:
        raise ValueError(
            f"Derived cash balance is negative ({cash_value:.2f}); check positions workbook totals and benchmark weight"
        )
    if abs(cash_value) < 1e-6:
        cash_value = 0.0

    existing_spy_row = workbook_active_meta.get("SPY") or any_meta.get("SPY") or {}
    snapshot_spy_rows = download_spy_history((snapshot_dt - timedelta(days=10)).isoformat(), (snapshot_dt + timedelta(days=2)).isoformat())
    snapshot_spy_dates = sorted(snapshot_spy_rows)
    snapshot_spy_date = nearest_available_date(snapshot_spy_dates, snapshot_date) if snapshot_spy_dates else None
    snapshot_previous_spy_date = (
        trading_sessions_back(snapshot_spy_dates, snapshot_spy_date, 5)
        if snapshot_spy_date and len(snapshot_spy_dates) > 1
        else None
    )
    reference_spy_price = float(snapshot_spy_rows.get(snapshot_spy_date) or existing_spy_row.get("buyPrice") or existing_spy_row.get("currentPrice") or 0)
    valuation_spy_price = float(snapshot_spy_rows.get(snapshot_spy_date) or existing_spy_row.get("currentPrice") or reference_spy_price or 0)
    previous_spy_price = float(snapshot_spy_rows.get(snapshot_previous_spy_date) or valuation_spy_price or reference_spy_price or 0)
    if reference_spy_price <= 0 or valuation_spy_price <= 0:
        reference_spy_price = valuation_spy_price = float(existing_summary.get("benchmarkCurrentPrice") or 1)
        previous_spy_price = previous_spy_price or valuation_spy_price

    benchmark_row, benchmark_shares = build_benchmark_row(
        active_meta,
        any_meta,
        snapshot_spy_date or snapshot_date,
        snapshot_spy_date or snapshot_date,
        reference_spy_price,
        valuation_spy_price,
        benchmark_value,
        previous_spy_price,
    )
    exited_db_rows = build_exited_rows(ledger_rows, scale_by_ticker, active_meta, any_meta)

    if balance_rows:
        spy_start_date = min(balance_rows[0]["date"], reference_date)
        spy_rows = download_spy_history(spy_start_date, TRACKER_DATE_CUTOFF.isoformat())
        if reference_date not in spy_rows:
            raise ValueError(f"Missing SPY close for {reference_date}")
        if valuation_date not in spy_rows:
            raise ValueError(f"Missing SPY close for {valuation_date}")
        daily_rows = build_daily_history(balance_rows, spy_rows)
        weekly_history = build_weekly_history(daily_rows)
        portfolio_returns = [row["portfolioReturn"] for row in weekly_history]
        benchmark_returns = [row["benchmarkReturn"] for row in weekly_history]
        latest_balance = balance_rows[-1]
        summary = {
            "snapshotDate": snapshot_date,
            "referenceDate": reference_date,
            "stockValue": stock_value,
            "stockPnl": sum(row[19] for row in active_db_rows),
            "benchmarkValue": benchmark_value,
            "cashValue": cash_value,
            "portfolioValue": portfolio_value,
            "portfolioStartValue": balance_rows[0]["portfolioValue"],
            "sinceStart": portfolio_value / balance_rows[0]["portfolioValue"] - 1 if balance_rows[0]["portfolioValue"] else 0,
            "benchmarkCurrentPrice": valuation_spy_price,
            "benchmarkShares": benchmark_shares,
            "portfolioVol": sample_std(portfolio_returns) * math.sqrt(52),
            "benchmarkVol": sample_std(benchmark_returns) * math.sqrt(52),
            "spyWeeklyReturn": weekly_history[-1]["benchmarkReturn"] if weekly_history else 0,
        }

        snapshot_row = next((row for row in daily_rows if row["date"] == latest_balance["date"]), None)
        if snapshot_row and round(snapshot_row["portfolioValue"], 2) != round(latest_balance["portfolioValue"], 2):
            raise ValueError(
                f"Snapshot total mismatch: expected {latest_balance['portfolioValue']:.2f}, got {snapshot_row['portfolioValue']:.2f}"
            )
    else:
        daily_rows = existing_daily_history
        weekly_history = existing_weekly_history
        portfolio_start_value = float(existing_summary.get("portfolioStartValue") or daily_rows[0]["portfolioValue"])
        summary = {
            "snapshotDate": snapshot_date,
            "referenceDate": reference_date,
            "stockValue": stock_value,
            "stockPnl": sum(row[19] for row in active_db_rows),
            "benchmarkValue": benchmark_value,
            "cashValue": cash_value,
            "portfolioValue": portfolio_value,
            "portfolioStartValue": portfolio_start_value,
            "sinceStart": portfolio_value / portfolio_start_value - 1 if portfolio_start_value else existing_summary.get("sinceStart", 0),
            "benchmarkCurrentPrice": valuation_spy_price,
            "benchmarkShares": benchmark_shares,
            "portfolioVol": float(existing_summary.get("portfolioVol") or 0),
            "benchmarkVol": float(existing_summary.get("benchmarkVol") or 0),
            "spyWeeklyReturn": float(existing_summary.get("spyWeeklyReturn") or 0),
        }

    payload = build_payload([benchmark_row, *active_db_rows], exited_db_rows, weekly_history, daily_rows, summary)
    write_module(output_path, payload)

    print(
        json.dumps(
            {
                "activeRows": len(payload["activeRows"]),
                "exitedRows": len(payload["exitedRows"]),
                "portfolioValue": round(summary["portfolioValue"], 2),
                "stockValue": round(summary["stockValue"], 2),
                "benchmarkValue": round(summary["benchmarkValue"], 2),
                "cashValue": round(summary["cashValue"], 2),
                "stockPnl": round(summary["stockPnl"], 2),
                "portfolioVol": round(summary["portfolioVol"], 4),
                "benchmarkVol": round(summary["benchmarkVol"], 4),
                "dailyRows": len(payload["dailyHistory"]),
                "weeklyRows": len(payload["weeklyHistory"]),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
