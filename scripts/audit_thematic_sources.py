#!/usr/bin/env python3

import argparse
import ast
import base64
import csv
import gzip
import json
import math
import operator
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


TRACKER_DATE_CUTOFF = date(2026, 3, 31)
THEME_NAME_MAP = {
    "AI Infra": "AI-Industrial",
    "Silver Economy": "Silver",
    "Batteries": "Battery",
    "Waste Management": "Waste",
    "Digital Finance": "Payments",
    "Legacy": "Legacy Software",
    "Water PFAs": "Water PFAS",
    "Securities": "Security",
}
DB_ROW_COLUMNS = [
    "id",
    "ticker",
    "company",
    "theme",
    "subTheme",
    "buyPrice",
    "currentPrice",
    "entryDate",
    "exitDate",
    "shares",
    "benchmarkWeight",
    "stopLossPct",
    "status",
    "notes",
    "marketBeta",
    "valueBeta",
    "momentumBeta",
    "weeklyReturn",
    "currentValue",
    "pnlFromExcel",
    "sellPrice",
    "costBasis",
    "sellTotal",
    "realizedPnl",
    "realizedPnlPct",
]
SAFE_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
SAFE_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def parse_args():
    parser = argparse.ArgumentParser(description="Cross-check thematic portfolio source files")
    parser.add_argument("--tracker", required=True, help="Path to Entry_Exit tracker workbook")
    parser.add_argument("--balance", required=True, help="Path to Portfolio Balance tracker workbook")
    parser.add_argument("--positions-csv", required=True, help="Path to current positions CSV snapshot")
    parser.add_argument("--history-csv", required=True, help="Path to broker account history CSV")
    parser.add_argument(
        "--seed-module",
        default=str(Path(__file__).resolve().parents[1] / "lib" / "thematicWorkbookData.js"),
        help="Path to generated thematic workbook module",
    )
    parser.add_argument("--snapshot-date", help="Override snapshot date in YYYY-MM-DD format")
    parser.add_argument("--value-threshold", type=float, default=100.0, help="Material value mismatch threshold")
    parser.add_argument("--qty-threshold", type=float, default=0.01, help="Material quantity mismatch threshold")
    parser.add_argument("--max-issues", type=int, default=12, help="Maximum issues to print in text mode")
    parser.add_argument("--output-json", help="Optional path to write the full JSON report")
    parser.add_argument("--format", choices=["text", "json"], default="text", help="Stdout format")
    return parser.parse_args()


def theme_name(raw):
    return THEME_NAME_MAP.get(raw, raw)


def normalize_theme(raw):
    return theme_name((raw or "").strip())


def normalize_ledger_date(value):
    if not hasattr(value, "date") and not isinstance(value, date):
        return None
    normalized = value.date() if hasattr(value, "date") else value
    while normalized > TRACKER_DATE_CUTOFF:
        normalized = normalized.replace(year=normalized.year - 1)
    return normalized


def normalize_balance_date(value, previous_date=None):
    if not hasattr(value, "date") and not isinstance(value, date):
        return None
    normalized = value.date() if hasattr(value, "date") else value
    while normalized > TRACKER_DATE_CUTOFF:
        normalized = normalized.replace(year=normalized.year - 1)
    while previous_date and normalized < previous_date and (previous_date - normalized).days > 180:
        normalized = normalized.replace(year=normalized.year + 1)
    return normalized


def eval_numeric_formula(node):
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.UnaryOp) and type(node.op) in SAFE_UNARY_OPERATORS:
        return SAFE_UNARY_OPERATORS[type(node.op)](eval_numeric_formula(node.operand))
    if isinstance(node, ast.BinOp) and type(node.op) in SAFE_BINARY_OPERATORS:
        return SAFE_BINARY_OPERATORS[type(node.op)](eval_numeric_formula(node.left), eval_numeric_formula(node.right))
    raise ValueError(f"Unsupported formula: {ast.dump(node)}")


def as_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if not stripped:
            return None
        if stripped.lower() in {"nan", "inf", "+inf", "-inf"}:
            return None
        if stripped.startswith("="):
            stripped = stripped[1:]
        try:
            numeric = float(stripped)
            return numeric if math.isfinite(numeric) else None
        except (TypeError, ValueError):
            try:
                parsed = ast.parse(stripped, mode="eval")
                numeric = float(eval_numeric_formula(parsed.body))
                return numeric if math.isfinite(numeric) else None
            except (SyntaxError, TypeError, ValueError, ZeroDivisionError):
                return None
    try:
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    except (TypeError, ValueError):
        return None


def parse_pct(value):
    if value is None:
        return None
    stripped = str(value).replace(",", "").strip()
    if not stripped or stripped.lower() == "nan":
        return None
    if stripped.endswith("%"):
        stripped = stripped[:-1]
    try:
        return float(stripped) / 100.0
    except ValueError:
        return None


def load_balance_rows(balance_path):
    wb = load_workbook(balance_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    previous_date = None
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw_date = values[0] if len(values) > 0 else None
        total_balance = values[1] if len(values) > 1 else None
        since_start = values[3] if len(values) > 3 else None
        normalized_date = normalize_balance_date(raw_date, previous_date)
        if normalized_date is None or total_balance is None:
            continue
        rows.append(
            {
                "date": normalized_date,
                "portfolioValue": float(total_balance),
                "sinceStart": as_float(since_start),
            }
        )
        previous_date = normalized_date
    return rows


def load_positions_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append(
                {
                    "symbol": row["Symbol"].strip(),
                    "subtheme": normalize_theme(row["Subtheme"]),
                    "weight": parse_pct(row["Weights"]),
                    "value": as_float(row["Current Value"]),
                }
            )
    return rows


def load_history_csv(path, snapshot_date):
    parsed_rows = []
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = None
        for row in reader:
            if row and row[0].strip() == "Run Date":
                header = row
                continue
            if header is None or not row or not row[0].strip():
                continue
            try:
                run_date = datetime.strptime(row[0].strip(), "%m/%d/%Y").date()
            except ValueError:
                continue
            if run_date > snapshot_date:
                continue
            parsed_rows.append(dict(zip(header, row)))

    aggregated = defaultdict(
        lambda: {"buyQty": 0.0, "sellQty": 0.0, "buyAmount": 0.0, "sellAmount": 0.0, "dates": set()}
    )
    for row in parsed_rows:
        symbol = row["Symbol"].strip()
        if not symbol:
            continue
        action = row["Action"].upper()
        quantity = abs(as_float(row["Quantity"]) or 0.0)
        amount = abs(as_float(row["Amount ($)"]) or 0.0)
        aggregated[symbol]["dates"].add(row["Run Date"])
        if "BOUGHT" in action:
            aggregated[symbol]["buyQty"] += quantity
            aggregated[symbol]["buyAmount"] += amount
        elif "SOLD" in action:
            aggregated[symbol]["sellQty"] += quantity
            aggregated[symbol]["sellAmount"] += amount
    return parsed_rows, aggregated


def load_tracker_snapshot(path, snapshot_date):
    wb = load_workbook(path, data_only=False, read_only=True)
    ledger_ws = wb["Sheet1"]
    history_ws = wb["Historical Data"]

    corrected_dates = []
    current_theme = None
    entries_by_ticker = defaultdict(list)
    exits_by_ticker = defaultdict(list)
    row_records = []

    for row_index, row in enumerate(ledger_ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            current_theme = normalize_theme(str(row[0]))
        ticker = row[1] if len(row) > 1 else None
        if not ticker:
            continue
        ticker = str(ticker).strip()

        def cell(idx):
            return row[idx] if len(row) > idx else None

        record = {
            "row": row_index,
            "ticker": ticker,
            "theme": current_theme,
            "entries": [],
            "exits": [],
        }

        for date_idx, qty_idx, price_idx, total_idx in ((2, 3, 4, 5), (6, 7, 8, 9), (10, 11, 12, 13)):
            raw_date = cell(date_idx)
            normalized_date = normalize_ledger_date(raw_date)
            qty = as_float(cell(qty_idx))
            price = as_float(cell(price_idx))
            total = as_float(cell(total_idx))
            if normalized_date and raw_date and hasattr(raw_date, "date") and raw_date.date() != normalized_date:
                corrected_dates.append(
                    {
                        "ticker": ticker,
                        "row": row_index,
                        "kind": "entry",
                        "rawDate": raw_date.date().isoformat(),
                        "normalizedDate": normalized_date.isoformat(),
                    }
                )
            if qty or price or total:
                entry = {
                    "date": normalized_date,
                    "qty": qty or 0.0,
                    "price": price or 0.0,
                    "total": total or 0.0,
                }
                record["entries"].append(entry)
                if normalized_date and normalized_date <= snapshot_date:
                    entries_by_ticker[ticker].append(entry)

        for date_idx, qty_idx, price_idx, total_idx in ((16, 17, 18, 19), (20, 21, 22, 23)):
            raw_date = cell(date_idx)
            normalized_date = normalize_ledger_date(raw_date)
            qty = as_float(cell(qty_idx))
            price = as_float(cell(price_idx))
            total = as_float(cell(total_idx))
            if normalized_date and raw_date and hasattr(raw_date, "date") and raw_date.date() != normalized_date:
                corrected_dates.append(
                    {
                        "ticker": ticker,
                        "row": row_index,
                        "kind": "exit",
                        "rawDate": raw_date.date().isoformat(),
                        "normalizedDate": normalized_date.isoformat(),
                    }
                )
            if qty or price or total:
                exit_row = {
                    "date": normalized_date,
                    "qty": qty or 0.0,
                    "price": price or 0.0,
                    "total": total or 0.0,
                }
                record["exits"].append(exit_row)
                if normalized_date and normalized_date <= snapshot_date:
                    exits_by_ticker[ticker].append(exit_row)

        row_records.append(record)

    headers = [cell.value for cell in history_ws[1]]
    raw_prices = {}
    for row in history_ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[0]
        if raw_date is None:
            continue
        row_date = raw_date.date() if hasattr(raw_date, "date") else raw_date
        price_map = {}
        for index, symbol in enumerate(headers[1:], start=1):
            if symbol:
                price = as_float(row[index])
                if price is not None:
                    price_map[str(symbol).strip()] = price
        raw_prices[row_date] = price_map

    if snapshot_date not in raw_prices:
        available = [row_date for row_date in raw_prices if row_date <= snapshot_date]
        if not available:
            raise ValueError(f"No Historical Data price row on or before {snapshot_date.isoformat()}")
        snapshot_price_date = max(available)
    else:
        snapshot_price_date = snapshot_date
    snapshot_prices = raw_prices[snapshot_price_date]

    aggregated = {}
    for ticker in sorted(set(entries_by_ticker) | set(exits_by_ticker)):
        buy_qty = sum(entry["qty"] for entry in entries_by_ticker[ticker])
        sell_qty = sum(entry["qty"] for entry in exits_by_ticker[ticker])
        net_qty = buy_qty - sell_qty
        if net_qty <= 1e-8:
            continue
        price = snapshot_prices.get(ticker)
        if price is None:
            continue
        themes = sorted({record["theme"] for record in row_records if record["ticker"] == ticker and record["theme"]})
        aggregated[ticker] = {
            "ticker": ticker,
            "themes": themes,
            "buyQty": round(buy_qty, 6),
            "sellQty": round(sell_qty, 6),
            "netQty": round(net_qty, 6),
            "price": round(price, 6),
            "value": round(net_qty * price, 2),
            "rows": [
                {
                    "row": record["row"],
                    "theme": record["theme"],
                    "entries": record["entries"],
                    "exits": record["exits"],
                }
                for record in row_records
                if record["ticker"] == ticker
            ],
        }

    tracker_aggregate = {}
    for ticker in sorted(set(entries_by_ticker) | set(exits_by_ticker)):
        tracker_aggregate[ticker] = {
            "buyQty": round(sum(entry["qty"] for entry in entries_by_ticker[ticker]), 6),
            "sellQty": round(sum(entry["qty"] for entry in exits_by_ticker[ticker]), 6),
            "themes": sorted({record["theme"] for record in row_records if record["ticker"] == ticker and record["theme"]}),
        }

    return {
        "snapshotPriceDate": snapshot_price_date,
        "prices": snapshot_prices,
        "positionsByTicker": aggregated,
        "trackerByTicker": tracker_aggregate,
        "correctedDates": corrected_dates,
    }


def load_seed_module(path):
    text = Path(path).read_text()
    match = re.search(r"const THEMATIC_WORKBOOK_B64 = \[(.*?)\]\.join", text, re.S)
    if not match:
        raise ValueError(f"Could not locate THEMATIC_WORKBOOK_B64 in {path}")
    encoded = "".join(re.findall(r'"([A-Za-z0-9+/=]+)"', match.group(1)))
    payload = json.loads(gzip.decompress(base64.b64decode(encoded)).decode())
    active_rows = [dict(zip(DB_ROW_COLUMNS, row)) for row in payload["activeRows"]]

    benchmark_value = 0.0
    stock_by_ticker = defaultdict(lambda: {"value": 0.0, "themes": set()})
    for row in active_rows:
        current_value = as_float(row["currentValue"]) or 0.0
        if row["theme"] == "Benchmark":
            benchmark_value += current_value
            continue
        ticker = row["ticker"]
        stock_by_ticker[ticker]["value"] += current_value
        stock_by_ticker[ticker]["themes"].add(row["theme"])

    return {
        "summary": payload["summary"],
        "stockByTicker": {
            ticker: {
                "value": round(meta["value"], 2),
                "themes": sorted(meta["themes"]),
            }
            for ticker, meta in stock_by_ticker.items()
        },
        "benchmarkValue": round(benchmark_value, 2),
    }


def build_positions_report(positions_rows, tracker_positions, threshold):
    positions_by_symbol = {row["symbol"]: row for row in positions_rows}
    common_numeric = [
        symbol
        for symbol in tracker_positions
        if symbol in positions_by_symbol and positions_by_symbol[symbol]["value"] is not None
    ]
    exact_matches = [
        symbol
        for symbol in common_numeric
        if abs(tracker_positions[symbol]["value"] - positions_by_symbol[symbol]["value"]) <= 0.01
    ]
    near_matches = [
        symbol
        for symbol in common_numeric
        if 0.01 < abs(tracker_positions[symbol]["value"] - positions_by_symbol[symbol]["value"]) <= threshold
    ]
    mismatches = []
    for symbol in common_numeric:
        diff = round(tracker_positions[symbol]["value"] - positions_by_symbol[symbol]["value"], 2)
        if abs(diff) > threshold:
            mismatches.append(
                {
                    "symbol": symbol,
                    "trackerValue": tracker_positions[symbol]["value"],
                    "positionsValue": positions_by_symbol[symbol]["value"],
                    "valueDiff": diff,
                    "trackerThemes": tracker_positions[symbol]["themes"],
                    "positionsTheme": positions_by_symbol[symbol]["subtheme"],
                }
            )
    mismatches.sort(key=lambda row: abs(row["valueDiff"]), reverse=True)

    missing_or_nan = []
    for symbol, tracker_row in tracker_positions.items():
        position_row = positions_by_symbol.get(symbol)
        if position_row is None or position_row["value"] is None:
            missing_or_nan.append({"symbol": symbol, "trackerValue": tracker_row["value"]})
    missing_or_nan.sort(key=lambda row: row["trackerValue"], reverse=True)

    return {
        "positionsCount": len(positions_rows),
        "positionsValue": round(sum(row["value"] or 0.0 for row in positions_rows), 2),
        "commonNumericCount": len(common_numeric),
        "exactMatchCount": len(exact_matches),
        "nearMatchCount": len(near_matches),
        "missingOrNan": missing_or_nan,
        "materialMismatches": mismatches,
    }


def build_history_report(history_by_symbol, tracker_by_symbol, qty_threshold):
    mismatches = []
    for symbol in sorted(set(history_by_symbol) | set(tracker_by_symbol)):
        history_row = history_by_symbol.get(symbol)
        tracker_row = tracker_by_symbol.get(symbol)
        history_buy = history_row["buyQty"] if history_row else None
        tracker_buy = tracker_row["buyQty"] if tracker_row else None
        history_sell = history_row["sellQty"] if history_row else None
        tracker_sell = tracker_row["sellQty"] if tracker_row else None
        buy_diff = None if history_row is None or tracker_row is None else round(tracker_buy - history_buy, 6)
        sell_diff = None if history_row is None or tracker_row is None else round(tracker_sell - history_sell, 6)
        if (
            history_row is None
            or tracker_row is None
            or abs(buy_diff or 0.0) > qty_threshold
            or abs(sell_diff or 0.0) > qty_threshold
        ):
            mismatches.append(
                {
                    "symbol": symbol,
                    "historyBuyQty": None if history_row is None else round(history_buy, 6),
                    "trackerBuyQty": None if tracker_row is None else round(tracker_buy, 6),
                    "buyQtyDiff": buy_diff,
                    "historySellQty": None if history_row is None else round(history_sell, 6),
                    "trackerSellQty": None if tracker_row is None else round(tracker_sell, 6),
                    "sellQtyDiff": sell_diff,
                    "trackerThemes": None if tracker_row is None else tracker_row["themes"],
                }
            )
    mismatches.sort(key=lambda row: max(abs(row["buyQtyDiff"] or 0.0), abs(row["sellQtyDiff"] or 0.0)), reverse=True)
    return {
        "historySymbolCount": len(history_by_symbol),
        "trackerSymbolCount": len(tracker_by_symbol),
        "materialMismatches": mismatches,
    }


def build_seed_report(seed_snapshot, tracker_positions, threshold):
    seed_positions = seed_snapshot["stockByTicker"]
    mismatches = []
    for symbol in sorted(set(seed_positions) | set(tracker_positions)):
        seed_row = seed_positions.get(symbol)
        tracker_row = tracker_positions.get(symbol)
        seed_value = seed_row["value"] if seed_row else None
        tracker_value = tracker_row["value"] if tracker_row else None
        value_diff = None if seed_row is None or tracker_row is None else round(tracker_value - seed_value, 2)
        if seed_row is None or tracker_row is None or abs(value_diff or 0.0) > threshold:
            mismatches.append(
                {
                    "symbol": symbol,
                    "trackerValue": tracker_value,
                    "seedValue": seed_value,
                    "valueDiff": value_diff,
                    "trackerThemes": None if tracker_row is None else tracker_row["themes"],
                    "seedThemes": None if seed_row is None else seed_row["themes"],
                }
            )
    mismatches.sort(key=lambda row: abs(row["valueDiff"] or 0.0), reverse=True)
    stock_value = round(sum(row["value"] for row in tracker_positions.values()), 2)
    return {
        "seedStockValue": round(sum(row["value"] for row in seed_positions.values()), 2),
        "trackerStockValue": stock_value,
        "seedBenchmarkValue": seed_snapshot["benchmarkValue"],
        "seedCashValue": round(as_float(seed_snapshot["summary"].get("cashValue")) or 0.0, 2),
        "seedPortfolioValue": round(as_float(seed_snapshot["summary"].get("portfolioValue")) or 0.0, 2),
        "stockValueDiff": round(stock_value - sum(row["value"] for row in seed_positions.values()), 2),
        "materialMismatches": mismatches,
    }


def build_report(args):
    balance_rows = load_balance_rows(args.balance)
    if not balance_rows:
        raise ValueError("Balance tracker did not contain any usable rows")
    snapshot_date = date.fromisoformat(args.snapshot_date) if args.snapshot_date else balance_rows[-1]["date"]

    positions_rows = load_positions_csv(args.positions_csv)
    history_rows, history_by_symbol = load_history_csv(args.history_csv, snapshot_date)
    tracker_snapshot = load_tracker_snapshot(args.tracker, snapshot_date)
    seed_snapshot = load_seed_module(args.seed_module) if args.seed_module and Path(args.seed_module).exists() else None

    report = {
        "snapshotDate": snapshot_date.isoformat(),
        "balance": {
            "startValue": round(balance_rows[0]["portfolioValue"], 2),
            "latestValue": round(balance_rows[-1]["portfolioValue"], 2),
            "sinceStart": round(balance_rows[-1]["sinceStart"] or 0.0, 10),
            "rowCount": len(balance_rows),
        },
        "positions": build_positions_report(positions_rows, tracker_snapshot["positionsByTicker"], args.value_threshold),
        "history": build_history_report(history_by_symbol, tracker_snapshot["trackerByTicker"], args.qty_threshold),
        "tracker": {
            "snapshotPriceDate": tracker_snapshot["snapshotPriceDate"].isoformat(),
            "stockValue": round(sum(row["value"] for row in tracker_snapshot["positionsByTicker"].values()), 2),
            "symbolCount": len(tracker_snapshot["positionsByTicker"]),
            "normalizedDates": tracker_snapshot["correctedDates"],
        },
        "historyRowsParsed": len(history_rows),
    }
    if seed_snapshot is not None:
        report["seed"] = build_seed_report(seed_snapshot, tracker_snapshot["positionsByTicker"], args.value_threshold)
    return report


def print_text_report(report, max_issues):
    print(f"Snapshot date: {report['snapshotDate']}")
    print(
        f"Balance tracker: start ${report['balance']['startValue']:,.2f} -> latest ${report['balance']['latestValue']:,.2f} "
        f"({report['balance']['sinceStart'] * 100:.4f}%)"
    )
    print(
        f"Tracker reconstruction: {report['tracker']['symbolCount']} symbols, stock value ${report['tracker']['stockValue']:,.2f}, "
        f"price row {report['tracker']['snapshotPriceDate']}"
    )
    print(
        f"Positions CSV: {report['positions']['positionsCount']} rows, value ${report['positions']['positionsValue']:,.2f}, "
        f"exact matches {report['positions']['exactMatchCount']}/{report['positions']['commonNumericCount']}"
    )
    if "seed" in report:
        print(
            f"Seed module: stock ${report['seed']['seedStockValue']:,.2f}, benchmark ${report['seed']['seedBenchmarkValue']:,.2f}, "
            f"cash ${report['seed']['seedCashValue']:,.2f}, portfolio ${report['seed']['seedPortfolioValue']:,.2f}"
        )

    if report["tracker"]["normalizedDates"]:
        print("Normalized tracker dates:")
        for row in report["tracker"]["normalizedDates"][:max_issues]:
            print(
                f"  - {row['ticker']} row {row['row']} {row['kind']} date {row['rawDate']} -> {row['normalizedDate']}"
            )

    if report["positions"]["missingOrNan"]:
        print("Positions CSV missing or NaN symbols:")
        for row in report["positions"]["missingOrNan"][:max_issues]:
            print(f"  - {row['symbol']}: tracker ${row['trackerValue']:,.2f}")

    if report["positions"]["materialMismatches"]:
        print("Material Positions CSV mismatches:")
        for row in report["positions"]["materialMismatches"][:max_issues]:
            print(
                f"  - {row['symbol']}: tracker ${row['trackerValue']:,.2f} vs positions ${row['positionsValue']:,.2f} "
                f"(diff ${row['valueDiff']:,.2f})"
            )

    if report["history"]["materialMismatches"]:
        print("Material history vs tracker quantity mismatches:")
        for row in report["history"]["materialMismatches"][:max_issues]:
            print(
                f"  - {row['symbol']}: buy diff {row['buyQtyDiff']} / sell diff {row['sellQtyDiff']}"
            )

    if "seed" in report and report["seed"]["materialMismatches"]:
        print("Material seed vs tracker mismatches:")
        for row in report["seed"]["materialMismatches"][:max_issues]:
            print(
                f"  - {row['symbol']}: tracker {row['trackerValue']} vs seed {row['seedValue']} "
                f"(diff {row['valueDiff']})"
            )


def main():
    args = parse_args()
    report = build_report(args)
    if args.output_json:
        Path(args.output_json).write_text(json.dumps(report, indent=2) + "\n")
    if args.format == "json":
        print(json.dumps(report, indent=2))
    else:
        print_text_report(report, args.max_issues)


if __name__ == "__main__":
    main()
